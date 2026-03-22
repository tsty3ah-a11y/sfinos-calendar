"""
Generate SQL to sync the client list from the Excel file.
- Match new clients to existing by name
- UPDATE matched clients (preserve notes + route_id)
- INSERT truly new clients
- DELETE old clients not in the new list
"""
import openpyxl
import re

EXCEL_PATH = "/Users/thanospangios/Downloads/ΝΤΙΝΟΣ NEW AREA 2026.xlsx"
OUTPUT_PATH = "/Users/thanospangios/Downloads/sfinos calendar/vet-planner/supabase_client_sync.sql"

# Skip rows that are headers/footers/blanks
SKIP_NAMES = {
    '', 'ΣΧΗΜΑΤΑΡΙ ΩΡΟΠΟΣ ΔΗΛΕΣΙ ΧΑΛΚΙΔΑ', 'ΝΕΟΙ ΠΕΛΑΤΕΣ', 'ΣΥΝΟΛΟ: 192',
}

# City-to-route mapping based on existing seed data
CITY_ROUTE = {
    # ΑΘΗΝΑ 1
    'ΑΓ. ΑΝΑΡΓΥΡΟΙ': 'ΑΘΗΝΑ 1',
    'ΑΧΑΡΝΑΙ': 'ΑΘΗΝΑ 1',
    'ΙΛΙΟΝ': 'ΑΘΗΝΑ 1',
    'ΚΑΜΑΤΕΡΟ': 'ΑΘΗΝΑ 1',
    'ΛΥΚΟΒΡΥΣΗ': 'ΑΘΗΝΑ 1',
    'ΜΕΤΑΜΟΡΦΩΣΗ': 'ΑΘΗΝΑ 1',
    'ΠΕΡΙΣΤΕΡΙ': 'ΑΘΗΝΑ 1',
    'ΠΕΡΙΣΤΕΡΙ - Ν.ΖΩΗ': 'ΑΘΗΝΑ 1',
    'ΠΕΤΡΟΥΠΟΛΗ': 'ΑΘΗΝΑ 1',
    # ΑΘΗΝΑ 2
    'ΑΓΙΑ ΒΑΡΒΑΡΑ': 'ΑΘΗΝΑ 2',
    'ΑΙΓΑΛΕΩ': 'ΑΘΗΝΑ 2',
    'ΒΡΙΛΗΣΣΙΑ': 'ΑΘΗΝΑ 2',
    'ΒΡΙΛΙΣΣΙΑ': 'ΑΘΗΝΑ 2',
    'ΚΟΡΥΔΑΛΛΟΣ': 'ΑΘΗΝΑ 2',
    'ΚΟΡΥΔΑΛΛΌΣ': 'ΑΘΗΝΑ 2',
    'ΜΕΛΙΣΣΙΑ': 'ΑΘΗΝΑ 2',
    'ΝΙΚΑΙΑ': 'ΑΘΗΝΑ 2',
    'Π. ΠΕΝΤΕΛΗ': 'ΑΘΗΝΑ 2',
    'ΠΕΥΚΗ': 'ΑΘΗΝΑ 2',
    'ΝΕΑ ΕΡΥΘΡΑΙΑ': 'ΑΘΗΝΑ 2',
    'Ν. ΠΕΝΤΕΛΗ': 'ΑΘΗΝΑ 2',
    'ΧΑΙΔΑΡΙ': 'ΑΘΗΝΑ 2',
    # ΑΘΗΝΑ 3
    'ΑΓ. ΙΩΑΝΝΗΣ ΡΕΝΤΗΣ': 'ΑΘΗΝΑ 3',
    'ΑΓ. ΣΟΦΙΑ ΠΕΙΡΑΙΑΣ': 'ΑΘΗΝΑ 3',
    'ΑΓΙΟΣ ΣΤΕΦΑΝΟΣ': 'ΑΘΗΝΑ 3',
    'ΔΡΑΠΕΤΣΩΝΑ': 'ΑΘΗΝΑ 3',
    'ΚΑΜΙΝΙΑ': 'ΑΘΗΝΑ 3',
    'ΚΗΦΙΣΙΑ': 'ΑΘΗΝΑ 3',
    'ΛΑΜΠΡΙΝΗ': 'ΑΘΗΝΑ 3',
    'Ν.ΚΗΦΙΣΙΑ': 'ΑΘΗΝΑ 3',
    'ΝΕΑ ΚΗΦΙΣΙΑ': 'ΑΘΗΝΑ 3',
    'Ν. ΦΑΛΗΡΟ': 'ΑΘΗΝΑ 3',
    'Π. ΦΑΛΗΡΟ': 'ΑΘΗΝΑ 3',
    'ΠΕΙΡΑΙΑΣ': 'ΑΘΗΝΑ 3',
    'ΠΕΙΡΑΙΑΣ-ΚΑΛΛΙΠΟΛΗ': 'ΑΘΗΝΑ 3',
    'ΠΕΙΡΑΙΚΗ': 'ΑΘΗΝΑ 3',
    'ΠΑΣΑΛΙΜΑΝΙ': 'ΑΘΗΝΑ 3',
    'ΜΑΡΟΥΣΙ': 'ΑΘΗΝΑ 3',
    'Ν. ΨΥΧΙΚΟ': 'ΑΘΗΝΑ 3',
    'Ν.ΨΥΧΙΚΟ': 'ΑΘΗΝΑ 3',
    'ΦΙΛΟΘΕΗ': 'ΑΘΗΝΑ 3',
    'ΧΑΛΑΝΔΡΙ': 'ΑΘΗΝΑ 3',
    'ΚΑΤΩ ΧΑΛΑΝΔΡΙ': 'ΑΘΗΝΑ 3',
    # ΑΘΗΝΑ 4
    'ΑΛΙΜΟΣ': 'ΑΘΗΝΑ 4',
    'ΑΝΟΙΞΗ': 'ΑΘΗΝΑ 4',
    'ΚΑΛΛΙΘΕΑ': 'ΑΘΗΝΑ 4',
    'ΚΑΛΛΙΘΈΑ': 'ΑΘΗΝΑ 4',
    'ΚΕΡΑΤΕΑ': 'ΑΘΗΝΑ 4',
    'ΚΕΡΑΤΣΙΝΙ': 'ΑΘΗΝΑ 4',
    'ΜΟΣΧΑΤΟ': 'ΑΘΗΝΑ 4',
    'Ν. ΗΡΑΚΛΕΙΟ': 'ΑΘΗΝΑ 4',
    'Ν.ΗΡΑΚΛΕΙΟ': 'ΑΘΗΝΑ 4',
    'ΤΑΥΡΟΣ': 'ΑΘΗΝΑ 4',
    # ΕΥΒΟΙΑ
    'ΣΧΗΜΑΤΑΡΙ': 'ΕΥΒΟΙΑ',
    'ΑΤΑΛΑΝΤΗ': 'ΕΥΒΟΙΑ',
    'ΝΕΑ ΠΑΛΑΤΙΑ ΩΡΟΠΟΥ': 'ΕΥΒΟΙΑ',
    'ΔΗΛΕΣΙ': 'ΕΥΒΟΙΑ',
    'ΔΡΟΣΙΑ ΧΑΛΚΙΔΑΣ': 'ΕΥΒΟΙΑ',
    'ΕΡΕΤΡΙΑ': 'ΕΥΒΟΙΑ',
    'ΙΣΤΙΑΙΑ': 'ΕΥΒΟΙΑ',
    'Ν. ΑΡΤΑΚΗ': 'ΕΥΒΟΙΑ',
    'ΧΑΛΚΙΔΑ': 'ΕΥΒΟΙΑ',
    'ΔΡΟΣΙΑ': 'ΕΥΒΟΙΑ',
    'ΚΑΠΑΝΔΡΙΤΙ': 'ΕΥΒΟΙΑ',
    'ΠΟΛΥΔΕΝΔΡΙ ΚΑΠΑΝΔΡΙΤΗ': 'ΕΥΒΟΙΑ',
    'ΡΟΔΟΠΟΛΗ': 'ΕΥΒΟΙΑ',
    'Ν.ΙΩΝΙΑ': 'ΕΥΒΟΙΑ',
    'ΝΕΑ ΦΙΛΑΔΕΛΦΕΙΑ': 'ΕΥΒΟΙΑ',
    'ΠΕΡΙΣΣΟΣ ΝΕΑ ΙΩΝΙΑ': 'ΕΥΒΟΙΑ',
    # New areas — best guess
    'ΑΝΩ ΛΙΟΣΙΑ': 'ΑΘΗΝΑ 1',
    'ΠΑΡΝΗΘΑ': 'ΑΘΗΝΑ 1',
    'ΜΕΝΙΔΙ': 'ΑΘΗΝΑ 1',
}

def escape_sql(s):
    if s is None:
        return 'NULL'
    s = str(s).strip()
    if s == '':
        return 'NULL'
    return "'" + s.replace("'", "''") + "'"

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Φύλλο1']

    new_clients = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6, values_only=True):
        name = str(row[0]).strip() if row[0] else ''
        if name in SKIP_NAMES or name == 'None':
            continue
        region = str(row[1]).strip() if row[1] else ''
        address = str(row[2]).strip() if row[2] else ''
        city = str(row[3]).strip() if row[3] else ''
        phone = str(row[4]).strip() if row[4] else ''
        mobile = str(row[5]).strip() if row[5] else ''
        # Clean up phone numbers - some come as integers
        if phone and phone.replace('.', '').replace('0', '').isdigit():
            phone = str(int(float(phone)))
        if mobile and mobile.replace('.', '').replace('0', '').isdigit():
            mobile = str(int(float(mobile)))
        # Fix known missing city
        if name == 'ΜΙΧΑΗΛ ΑΝΝΕΤΑ' and not city:
            city = 'ΚΑΠΑΝΔΡΙΤΙ'
        new_clients.append({
            'name': name, 'region': region, 'address': address,
            'city': city, 'phone': phone, 'mobile': mobile,
        })

    print(f"Parsed {len(new_clients)} clients from Excel")

    # Build the name list for matching
    name_list = [c['name'] for c in new_clients]

    # Generate SQL
    lines = []
    lines.append("-- ══════════════════════════════════════════════════════════════")
    lines.append("-- VetPlan — Client Sync from ΝΤΙΝΟΣ NEW AREA 2026.xlsx")
    lines.append(f"-- {len(new_clients)} clients total")
    lines.append("-- Run this in Supabase SQL Editor")
    lines.append("-- ══════════════════════════════════════════════════════════════")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # Step 1: Delete clients NOT in the new list
    # Build a list of all names for the WHERE NOT IN clause
    lines.append("-- ═══ STEP 1: Delete clients not in new list ═══")
    lines.append("-- First delete related visits and appointments, then clients")
    lines.append("")

    # Build escaped name list
    escaped_names = ", ".join(escape_sql(n) for n in name_list)

    lines.append("-- Delete visits for clients being removed")
    lines.append(f"DELETE FROM visits WHERE client_id IN (")
    lines.append(f"  SELECT id FROM clients WHERE name NOT IN ({escaped_names})")
    lines.append(f");")
    lines.append("")

    lines.append("-- Delete appointments for clients being removed")
    lines.append(f"DELETE FROM appointments WHERE client_id IN (")
    lines.append(f"  SELECT id FROM clients WHERE name NOT IN ({escaped_names})")
    lines.append(f");")
    lines.append("")

    lines.append("-- Delete the clients themselves")
    lines.append(f"DELETE FROM clients WHERE name NOT IN ({escaped_names});")
    lines.append("")

    # Step 2: Update existing clients with new details (preserve notes & route_id)
    lines.append("-- ═══ STEP 2: Update existing clients (preserve notes & route_id) ═══")
    lines.append("")
    for c in new_clients:
        lines.append(f"UPDATE clients SET")
        updates = []
        if c['region']:
            updates.append(f"  region = {escape_sql(c['region'])}")
        if c['address']:
            updates.append(f"  address = {escape_sql(c['address'])}")
        if c['city']:
            updates.append(f"  city = {escape_sql(c['city'])}")
        if c['phone']:
            updates.append(f"  phone = {escape_sql(c['phone'])}")
        if c['mobile']:
            updates.append(f"  mobile = {escape_sql(c['mobile'])}")
        updates.append(f"  is_active = true")
        lines.append(",\n".join(updates))
        lines.append(f"WHERE name = {escape_sql(c['name'])};")
        lines.append("")

    # Step 3: Insert truly new clients (ones that don't exist by name)
    lines.append("-- ═══ STEP 3: Insert new clients that don't exist yet ═══")
    lines.append("")

    for c in new_clients:
        route = CITY_ROUTE.get(c['city'], 'ΑΘΗΝΑ 1')  # fallback
        lines.append(f"INSERT INTO clients (route_id, name, region, address, city, phone, mobile)")
        lines.append(f"SELECT r.id, {escape_sql(c['name'])}, {escape_sql(c['region'])}, {escape_sql(c['address'])}, {escape_sql(c['city'])}, {escape_sql(c['phone'])}, {escape_sql(c['mobile'])}")
        lines.append(f"FROM routes r WHERE r.name = '{route}'")
        lines.append(f"AND NOT EXISTS (SELECT 1 FROM clients WHERE name = {escape_sql(c['name'])});")
        lines.append("")

    lines.append("COMMIT;")
    lines.append("")
    lines.append("-- ═══ Verify ═══")
    lines.append("-- SELECT count(*) FROM clients;  -- Should be ~192")
    lines.append("-- SELECT r.name, count(c.id) FROM routes r LEFT JOIN clients c ON c.route_id = r.id GROUP BY r.name ORDER BY r.name;")

    sql = "\n".join(lines)
    with open(OUTPUT_PATH, 'w') as f:
        f.write(sql)
    print(f"SQL written to {OUTPUT_PATH}")
    print(f"Client count: {len(new_clients)}")

    # Show new clients (not in seed)
    from collections import Counter
    route_counts = Counter()
    for c in new_clients:
        route = CITY_ROUTE.get(c['city'], 'ΑΘΗΝΑ 1 (guess)')
        route_counts[route] += 1
    print("\nRoute distribution:")
    for route, count in sorted(route_counts.items()):
        print(f"  {route}: {count}")

if __name__ == "__main__":
    main()
